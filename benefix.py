# PDF -> Excel scaper
# Author: Stefan Stancu

import subprocess
import glob
import re
import openpyxl

from page import Page

def check_dependencies():
    try:
        x = subprocess.Popen(["gs", "--version"], stdout=subprocess.PIPE)
    except OSError as e:
        if e.errno == 2:
            print("Error: GhostScript not installed or gs command not available")

def read_pdf(file_name):
    raw_text = []
    cmd = ["gs", "-sDEVICE=txtwrite", "-dNOPAUSE", "-dBATCH", "-sOutputFile=-", file_name]

    x = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    text, err = x.communicate()
    if not err:
        print(err)
        exit()

    for page in text.split("Page"):
        raw_text.append(re.split("\r\n |\n ", page)[1:21])

    raw_text.pop(0)
    return raw_text

def dump_page(ws, page):
    """
    Appends the page (row) to the sheet
    @param ws an open worksheet
    @page a page object
    """
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=str(page.start_date) + " UTC")
    ws.cell(row=row, column=2, value=str(page.end_date) + " UTC")
    ws.cell(row=row, column=3, value=page.product_name)
    ws.cell(row=row, column=4, value=page.state)
    ws.cell(row=row, column=5, value=page.rating)
    
    data = page.data_list()
    for column in range(6, 6 + len(data)):
        ws.cell(row=row, column=column, value=data[column - 6])


if __name__=="__main__":
    excel_file = "BeneFix Small Group Plans upload template.xlsx"

    check_dependencies()

    wb = openpyxl.load_workbook(excel_file)

    for file_name in glob.glob("*.pdf"):
        print("Parsing " + file_name)
        for raw_page_text in read_pdf(file_name):
            page = Page(raw_page_text)
            dump_page(wb.active, page)
    
    wb.save(excel_file)
